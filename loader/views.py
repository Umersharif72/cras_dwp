import os
import csv
import openpyxl
import xlrd
from django.db import connection
from django.conf import settings
from django.shortcuts import render, redirect
from .forms import CSVUploadForm
from langdetect import detect
from googletrans import Translator

translator = Translator()


def handle_uploaded_file(f):
    # Ensure the media directory exists
    if not os.path.exists(settings.MEDIA_ROOT):
        os.makedirs(settings.MEDIA_ROOT)

    # Use a safe filename function to handle special characters
    # This may include normalizing Unicode characters or escaping special characters
    safe_filename = f.name.encode('ascii', 'ignore').decode('ascii')  # Fallback to ASCII

    # Create the full path where the file will be saved
    file_path = os.path.join(settings.MEDIA_ROOT, safe_filename)
    
    try:
        # Open the file in binary write mode and save the uploaded file
        with open(file_path, 'wb+') as destination:
            for chunk in f.chunks():
                destination.write(chunk)
    except IOError as e:
        # Handle any I/O errors (e.g., permission issues, disk full, etc.)
        print(f"Error saving file {file_path}: {e}")
        raise

    return file_path



def upload_csv(request):
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_path = handle_uploaded_file(request.FILES['file'])
            table_name, _ = os.path.splitext(request.FILES['file'].name)
            create_table_and_insert_data(file_path, table_name)
            # Pass a context variable indicating upload success
            return render(request, 'main/landingpage.html', {'form': form, 'upload_success': True})
    else:
        form = CSVUploadForm()
    return render(request, 'main/landingpage.html', {'form': form})

def upload_success(request):
    return render(request, 'main/landingpage.html', {'upload_success': True})

def infer_data_type(value):
    if isinstance(value, int):
        return 'INT'
    elif isinstance(value, float):
        return 'FLOAT'
    elif isinstance(value, bool):
        return 'BIT'
    elif isinstance(value, str):
        return 'NVARCHAR(MAX)'  # Unicode text
    else:
        return 'NVARCHAR(MAX)'

def make_unique_headers(headers):
    seen = set()
    unique_headers = []
    for header in headers:
        original_header = header if header else 'Unnamed'
        new_header = original_header
        index = 1
        while new_header in seen:
            new_header = f"{original_header}_{index}"
            index += 1
        seen.add(new_header)
        unique_headers.append(new_header)
    return unique_headers
import chardet

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read(10000))  # Read a portion of the file to guess encoding
    return result['encoding']

def read_file_with_encoding(file_path):
    encoding = detect_encoding(file_path)
    try:
        with open(file_path, newline='', encoding=encoding) as file:
            return file.read()
    except UnicodeDecodeError:
        print(f"Error reading file with encoding {encoding}, trying 'latin1'")
        with open(file_path, newline='', encoding='latin1') as file:
            return file.read()

def create_table_and_insert_data(file_path, table_name):
    extension = os.path.splitext(file_path)[1].lower()
    table_name = connection.ops.quote_name(table_name)

    if extension == '.csv':
        encoding = detect_encoding(file_path)
        with open(file_path, newline='', encoding=encoding) as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)
            headers = make_unique_headers(headers)
            first_row = next(reader)

            data_types = [infer_data_type(value) for value in first_row]
            escaped_headers = [f"{connection.ops.quote_name(header)} {data_type}" for header, data_type in zip(headers, data_types)]

            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = %s", [table_name.strip('"')])
                table_exists = cursor.fetchone()[0] == 1

                if not table_exists:
                    create_table_sql = f"CREATE TABLE {table_name} ({', '.join(escaped_headers)})"
                    cursor.execute(create_table_sql)

                insert_sql = f"INSERT INTO {table_name} ({', '.join([connection.ops.quote_name(header) for header in headers])}) VALUES ({', '.join(['%s' for _ in headers])})"
                cursor.execute(insert_sql, first_row)

                for row in reader:
                    cursor.execute(insert_sql, row)
        print(f"Data inserted into table {table_name} successfully.")

    elif extension == '.xlsx':
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value if cell.value else 'Unnamed' for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = make_unique_headers(headers)
        first_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        data = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2)]

        if len(headers) == 0:
            print("No headers found in the file.")
            return

        data_types = [infer_data_type(value) for value in first_row]
        escaped_headers = [f"{connection.ops.quote_name(header)} {data_type}" for header, data_type in zip(headers, data_types)]

        if len(data) > 0:
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = %s", [table_name.strip('"')])
                table_exists = cursor.fetchone()[0] == 1

                if not table_exists:
                    create_table_sql = f"CREATE TABLE {table_name} ({', '.join(escaped_headers)})"
                    cursor.execute(create_table_sql)

                insert_sql = f"INSERT INTO {table_name} ({', '.join([connection.ops.quote_name(header) for header in headers])}) VALUES ({', '.join(['%s' for _ in headers])})"

                for row in data:
                    cursor.execute(insert_sql, row)

            print(f"Data inserted into table {table_name} successfully.")
        else:
            print("No data found in the file.")
        workbook.close()

    elif extension == '.xls':
        workbook = xlrd.open_workbook(file_path, encoding_override="utf-8")
        sheet = workbook.sheet_by_index(0)
        headers = sheet.row_values(0)
        headers = make_unique_headers(headers)
        first_row = sheet.row_values(1)
        data = [sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows)]

        if len(headers) == 0:
            print("No headers found in the file.")
            return

        data_types = [infer_data_type(value) for value in first_row]
        escaped_headers = [f"{connection.ops.quote_name(header)} {data_type}" for header, data_type in zip(headers, data_types)]

        if len(data) > 0:
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = %s", [table_name.strip('"')])
                table_exists = cursor.fetchone()[0] == 1

                if not table_exists:
                    create_table_sql = f"CREATE TABLE {table_name} ({', '.join(escaped_headers)})"
                    cursor.execute(create_table_sql)

                insert_sql = f"INSERT INTO {table_name} ({', '.join([connection.ops.quote_name(header) for header in headers])}) VALUES ({', '.join(['%s' for _ in headers])})"

                for row in data:
                    cursor.execute(insert_sql, row)

            print(f"Data inserted into table {table_name} successfully.")
        else:
            print("No data found in the file.")
        workbook.release_resources()

    else:
        print("Unsupported file format.")